"""
Tests for _upsert_xlsx_to_sqlite() in app2.py.

Covers normalized vs original XLSX detection, CSI regex filtering,
UPSERT logic, category derivation, and edge cases.
"""

import os
import sys
import pytest
from unittest.mock import patch

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def temp_modules_db(tmp_path):
    """Use a temporary SQLite DB for each test."""
    import agent.db as agent_db

    db_path = str(tmp_path / "test_classifications.db")
    with patch("agent.db._DB_PATH", db_path):
        if hasattr(agent_db._local, "conn") and agent_db._local.conn is not None:
            try:
                agent_db._local.conn.close()
            except Exception:
                pass
        agent_db._local.__dict__.clear()
        from agent.db import init_db
        init_db()
        yield db_path
        if hasattr(agent_db._local, "conn") and agent_db._local.conn is not None:
            try:
                agent_db._local.conn.close()
            except Exception:
                pass
        agent_db._local.__dict__.clear()


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def make_workbook(path, rows_by_sheet):
    """Write an XLSX at *path* with the given {sheet_name: [row_tuples]}."""
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in rows_by_sheet.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(list(row))
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# Normalized format (header row: code | name | description)
# ---------------------------------------------------------------------------

class TestNormalizedFormat:

    def test_normalized_header_detected(self, tmp_path, temp_modules_db):
        """First row col A = "code" triggers normalized mode; data row is upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A0", "Type S Mortar", "Standard mortar"),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 1

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT * FROM classifications WHERE code = '04 05 13.A0'"
        ).fetchone()
        assert row is not None
        assert row["name"] == "Type S Mortar"

    def test_normalized_header_case_insensitive(self, tmp_path, temp_modules_db):
        """'Code' and 'CODE' are also recognised as the normalized header."""
        from app2 import _upsert_xlsx_to_sqlite
        import openpyxl

        for i, header in enumerate(["Code", "CODE"]):
            xlsx_path = str(tmp_path / f"test_{i}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([header, "Name", "Desc"])
            ws.append([f"04 05 {10 + i:02d}.A0", f"Mortar {i}", "desc"])
            wb.save(xlsx_path)
            count = _upsert_xlsx_to_sqlite(xlsx_path)
            assert count == 1, f"Expected 1 for header {header!r}, got {count}"

    def test_normalized_skips_header_row(self, tmp_path, temp_modules_db):
        """The header row itself must NOT be inserted into SQLite."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A0", "Type S Mortar", "Mortar"),
        ]})
        _upsert_xlsx_to_sqlite(path)

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT * FROM classifications WHERE code = 'code'"
        ).fetchone()
        assert row is None

    def test_normalized_upserts_all_rows(self, tmp_path, temp_modules_db):
        """All three data rows are upserted and the return count matches."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A0", "Row One", "Desc 1"),
            ("04 05 13.A1", "Row Two", "Desc 2"),
            ("04 05 13.A2", "Row Three", "Desc 3"),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 3

    def test_normalized_description_optional(self, tmp_path, temp_modules_db):
        """Col C empty → description stored as '' and row is still upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A0", "Type S Mortar", None),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 1

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT description FROM classifications WHERE code = '04 05 13.A0'"
        ).fetchone()
        assert row is not None
        assert (row["description"] or "") == ""


# ---------------------------------------------------------------------------
# Original format (CSI code in col B)
# ---------------------------------------------------------------------------

class TestOriginalFormat:

    def test_original_csi_code_matched(self, tmp_path, temp_modules_db):
        """Col B matching '04 05 13' (bare section code) is upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("", "04 05 13", "Mortar"),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 1

    def test_original_csi_subcode_matched(self, tmp_path, temp_modules_db):
        """Col B = '04 05 13.A1' (sub-code with dot) is also upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("", "04 05 13.A1", "Type N Mortar"),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 1

    def test_original_non_csi_skipped(self, tmp_path, temp_modules_db):
        """Col B = 'Master Notes' (no leading digits) is NOT upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("", "Master Notes", "Some text"),
        ]})
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 0

    def test_original_short_row_skipped(self, tmp_path, temp_modules_db):
        """Row with fewer than 2 cells is skipped without raising an exception."""
        from app2 import _upsert_xlsx_to_sqlite
        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.append(["Only one cell"])
        path = str(tmp_path / "short.xlsx")
        wb.save(path)

        count = _upsert_xlsx_to_sqlite(path)
        assert count == 0


# ---------------------------------------------------------------------------
# Category derivation and UPSERT semantics
# ---------------------------------------------------------------------------

class TestCategoryAndUpsert:

    def test_category_derived_from_dot(self, tmp_path, temp_modules_db):
        """Code '04 05 13.A1' → category stored as '04 05 13'."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A1", "Type N Mortar", "Normal mortar"),
        ]})
        _upsert_xlsx_to_sqlite(path)

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT category FROM classifications WHERE code = '04 05 13.A1'"
        ).fetchone()
        assert row["category"] == "04 05 13"

    def test_category_empty_if_no_dot(self, tmp_path, temp_modules_db):
        """Code '04 05 13' (no dot) → category stored as ''."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13", "Mortar Base", "Base code"),
        ]})
        _upsert_xlsx_to_sqlite(path)

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT category FROM classifications WHERE code = '04 05 13'"
        ).fetchone()
        assert row["category"] == ""

    def test_upsert_updates_existing(self, tmp_path, temp_modules_db):
        """Same code upserted twice → latest name wins, only one row exists."""
        from app2 import _upsert_xlsx_to_sqlite
        import openpyxl

        def write_xlsx(filename, name):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["code", "name", "description"])
            ws.append(["04 05 13.A0", name, "desc"])
            p = str(tmp_path / filename)
            wb.save(p)
            return p

        _upsert_xlsx_to_sqlite(write_xlsx("first.xlsx", "Original Name"))
        _upsert_xlsx_to_sqlite(write_xlsx("second.xlsx", "Updated Name"))

        import agent.db as agent_db
        rows = agent_db.get_db().execute(
            "SELECT name FROM classifications WHERE code = '04 05 13.A0'"
        ).fetchall()
        assert len(rows) == 1
        assert rows[0]["name"] == "Updated Name"

    def test_upsert_source_set_to_xlsx(self, tmp_path, temp_modules_db):
        """All rows ingested from XLSX have source = 'xlsx'."""
        from app2 import _upsert_xlsx_to_sqlite

        path = make_workbook(tmp_path / "t.xlsx", {"Sheet1": [
            ("code", "name", "description"),
            ("04 05 13.A0", "Type S Mortar", "Mortar"),
        ]})
        _upsert_xlsx_to_sqlite(path)

        import agent.db as agent_db
        row = agent_db.get_db().execute(
            "SELECT source FROM classifications WHERE code = '04 05 13.A0'"
        ).fetchone()
        assert row["source"] == "xlsx"


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:

    def test_missing_file_returns_zero(self, tmp_path, temp_modules_db):
        """Non-existent path → returns 0, no exception raised."""
        from app2 import _upsert_xlsx_to_sqlite

        count = _upsert_xlsx_to_sqlite(str(tmp_path / "nonexistent.xlsx"))
        assert count == 0

    def test_empty_workbook_returns_zero(self, tmp_path, temp_modules_db):
        """XLSX with no data rows → returns 0."""
        from app2 import _upsert_xlsx_to_sqlite
        import openpyxl

        wb = openpyxl.Workbook()
        path = str(tmp_path / "empty.xlsx")
        wb.save(path)

        count = _upsert_xlsx_to_sqlite(path)
        assert count == 0


# ---------------------------------------------------------------------------
# 03d-Families parser
# ---------------------------------------------------------------------------

class TestFamiliesParser:
    """Tests for _parse_families: full-row reading and filtering."""

    def _make_families_workbook(self, path, data_rows):
        """Build a 03d-Families XLSX: 4 header rows then data_rows."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "03d-Families"
        # 4 header rows (rows 0-3 in 0-indexed, i.e. Excel rows 1-4)
        for _ in range(4):
            ws.append(["", "", "", "", ""])
        for row in data_rows:
            ws.append(list(row))
        wb.save(str(path))
        return str(path)

    def test_families_reads_beyond_row_10(self, tmp_path, temp_modules_db):
        """All 12 data rows (including rows 11-12) are returned — not just first 7."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("CASEWORK", "BASE", "4DRAWER", "SYN", "CASE_BASE_4DRAWER_SYN"),
            ("CASEWORK", "BASE", "6DRAWER", "SYN", "CASE_BASE_6DRAWER_SYN"),
            ("CASEWORK", "BASE", "2DRAWER", "SYN", "CASE_BASE_2DRAWER_SYN"),
            ("DOORS", "EXT", "SINGLE", "SYN", "DR_EXT_SINGLE_SYN"),
            ("DOORS", "EXT", "DOUBLE", "SYN", "DR_EXT_DOUBLE_SYN"),
            ("DOORS", "INT", "SINGLE", "SYN", "DR_INT_SINGLE_SYN"),
            ("DOORS", "INT", "DOUBLE", "SYN", "DR_INT_DOUBLE_SYN"),
            ("WINDOWS", "FIXED", "SINGLE", "SYN", "WIN_FIXED_SINGLE_SYN"),
            ("WINDOWS", "FIXED", "DOUBLE", "SYN", "WIN_FIXED_DOUBLE_SYN"),
            ("WINDOWS", "OPERABLE", "SINGLE", "SYN", "WIN_OP_SINGLE_SYN"),
            ("WINDOWS", "OPERABLE", "DOUBLE", "SYN", "WIN_OP_DOUBLE_SYN"),
            ("WALLS", "EXTERIOR", "PANEL", "SYN", "WL_EXT_PANEL_SYN"),
        ]
        path = self._make_families_workbook(tmp_path / "fam.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 12

    def test_families_skips_na_rows(self, tmp_path, temp_modules_db):
        """Rows where col E = '#N/A' are excluded from results."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("CASEWORK", "BASE", "4DRAWER", "SYN", "CASE_BASE_4DRAWER_SYN"),
            ("DOORS", "EXT", "SINGLE", "SYN", "#N/A"),
            ("WINDOWS", "FIXED", "DOUBLE", "SYN", "WIN_FIXED_DOUBLE_SYN"),
        ]
        path = self._make_families_workbook(tmp_path / "fam.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 2

    def test_families_skips_empty_category(self, tmp_path, temp_modules_db):
        """Rows where col A (category) is empty are excluded."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("CASEWORK", "BASE", "4DRAWER", "SYN", "CASE_BASE_4DRAWER_SYN"),
            ("", "EXT", "SINGLE", "SYN", "DR_EXT_SINGLE_SYN"),
            ("WINDOWS", "FIXED", "DOUBLE", "SYN", "WIN_FIXED_DOUBLE_SYN"),
        ]
        path = self._make_families_workbook(tmp_path / "fam.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 2


# ---------------------------------------------------------------------------
# 03e-Detail Name parser
# ---------------------------------------------------------------------------

class TestDetailNameParser:
    """Tests for _parse_detail_name: column positions, full-row reading, filtering."""

    def _make_detail_workbook(self, path, data_rows, sheet_name="03e-Detail Name "):
        """Build a 03e-Detail Name XLSX: 13 header rows then data_rows.

        Column layout (0-indexed): 9=category, 10=type, 11=adjective, 12=company, 13=name.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        empty = [""] * 14
        for _ in range(13):
            ws.append(empty)
        for row in data_rows:
            r = [""] * 14
            r[9] = row[0]   # category
            r[10] = row[1]  # type/function
            r[11] = row[2]  # adjective
            r[12] = row[3]  # company
            r[13] = row[4]  # generated detail name
            ws.append(r)
        wb.save(str(path))
        return str(path)

    def test_detail_name_parses_data_rows(self, tmp_path, temp_modules_db):
        """All 5 data rows from 03e-Detail Name are upserted."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("DETAIL ITEMS-SMART", "WALLS", "1HR_NLB", "SYN", "DETS_WALL_1HR_NLB_SYN"),
            ("DETAIL ITEMS-SMART", "WALLS", "GWB_6_ABOVE_CEILING", "SYN", "DETS_WALL_GWB_6_ABOVE_CEILING_SYN"),
            ("DETAIL ITEMS-SMART", "CEILINGS", "WALL_ANGLE", "SYN", "DETS_CLNG_WALL_ANGLE_SYN"),
            ("DETAIL ITEMS-SMART", "WALLS", "HOLLOW_MTL_FRAME", "SYN", "DETS_WALL_HOLLOW_MTL_FRAME_SYN"),
            ("DETAIL ITEMS-DUMB", "WALLS", "2HR_NLB", "SYN", "DETD_WALL_2HR_NLB_SYN"),
        ]
        path = self._make_detail_workbook(tmp_path / "det.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 5

    def test_detail_name_skips_na_rows(self, tmp_path, temp_modules_db):
        """Rows where col N (idx 13) = '#N/A' are excluded."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("DETAIL ITEMS-SMART", "WALLS", "1HR_NLB", "SYN", "DETS_WALL_1HR_NLB_SYN"),
            ("DETAIL ITEMS-SMART", "WALLS", "", "SYN", "#N/A"),
            ("DETAIL ITEMS-DUMB", "WALLS", "2HR_NLB", "SYN", "DETD_WALL_2HR_NLB_SYN"),
        ]
        path = self._make_detail_workbook(tmp_path / "det.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 2

    def test_detail_name_skips_empty_category(self, tmp_path, temp_modules_db):
        """Rows where col J (idx 9) is empty are excluded."""
        from app2 import _upsert_xlsx_to_sqlite

        data_rows = [
            ("DETAIL ITEMS-SMART", "WALLS", "1HR_NLB", "SYN", "DETS_WALL_1HR_NLB_SYN"),
            ("", "WALLS", "2HR_NLB", "SYN", "DETD_WALL_2HR_NLB_SYN"),
            ("DETAIL ITEMS-DUMB", "WALLS", "3HR_NLB", "SYN", "DETD_WALL_3HR_NLB_SYN"),
        ]
        path = self._make_detail_workbook(tmp_path / "det.xlsx", data_rows)
        count = _upsert_xlsx_to_sqlite(path)
        assert count == 2

    def test_detail_name_sheet_with_trailing_space(self, tmp_path, temp_modules_db):
        """Sheet named '03e-Detail Name ' (trailing space) is matched correctly."""
        from app2 import _upsert_xlsx_to_sqlite
        import agent.db as agent_db

        data_rows = [
            ("DETAIL ITEMS-SMART", "WALLS", "1HR_NLB", "SYN", "DETS_WALL_1HR_NLB_SYN"),
        ]
        path = self._make_detail_workbook(tmp_path / "det.xlsx", data_rows, sheet_name="03e-Detail Name ")
        _upsert_xlsx_to_sqlite(path)

        row = agent_db.get_db().execute(
            "SELECT * FROM classifications WHERE code = 'DETS_WALL_1HR_NLB_SYN'"
        ).fetchone()
        assert row is not None
        assert row["name"] == "DETAIL ITEMS-SMART"
