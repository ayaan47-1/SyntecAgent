"""Tests for the CCN rules engine (Step 2a: structural + referential rules).

Fixtures are built in-memory with openpyxl so the real workbook is never
committed. The synthetic ``05-Views`` sheet reproduces the real header layout
(columns detected by header text, not position) and seeds every hazard the
rules must catch: broken formula results, disagreeing generated-name columns,
duplicate names, empty composite fields, and phase abbreviations that do not
resolve to the sheet-17 Construction Nature vocabulary.
"""
import openpyxl
import pytest

from agent.ccn import views as ccn_views


# --- sheet-17 vocab blocks the referential rules resolve against ---------
_BLOCKS = [
    ("H", None, "LEVELS", ("Levels", None),
     [("04", None), ("05", None)]),
    ("J", "K", "CONSTRUCTION NATURE", ("Nature", "Nature-2"),
     [("NEW_New", "NEW"), ("EXT_Existing", "EXT"),
      ("EXD_Existing-Demolition", "EXD")]),
    ("M", "N", "DISCIPLINES", ("Discipline", "Disc-2"),
     [("ar-Architecture", "ar"), ("fa_Facility Analysis", "fa")]),
]

# Views columns, laid out with the *real* header labels so the reader must
# resolve them by text. Order is deliberately not the natural field order.
_VIEWS_HEADERS = {
    "A": "View Type",
    "B": "Abbreviation_View Type",
    "C": "Level",
    "D": "Abbreviation_Phase Name",
    "E": "Phase",
    "F": "Abbeviation_Discipline Name",   # sic — matches the workbook typo
    "G": "Discipline",
    "H": "View Name, Generated",
    "I": "View Name Value Only",
}

# Each tuple: (B view-type, C level, D phase, F discipline, H generated, I value-only)
_VIEWS_ROWS = [
    # 1 clean row — everything resolves, columns agree
    ("FLR-PLN_Floor Plan", "04", "new_New", "ar-Architecture",
     "FLR-PLN_04_new_ar", "FLR-PLN_04_new_ar"),
    # 2 phase 'ex' does not resolve (should be EXT) -> referential orphan
    ("FLR-PLN_Floor Plan", "04", "ex_Existing", "ar-Architecture",
     "FLR-PLN_04_ex_ar", "FLR-PLN_04_ex_ar"),
    # 3 another 'ex' orphan (grouped with row 2 by value)
    ("REF-CLG-PLN_Ceiling", "05", "ex_Existing", "ar-Architecture",
     "REF-CLG-PLN_05_ex_ar", "REF-CLG-PLN_05_ex_ar"),
    # 4 broken formula in the generated-name column -> structural
    ("FLR-PLN_Floor Plan", "04", "new_New", "ar-Architecture",
     "#REF!", "#REF!"),
    # 5 duplicate of row 1's generated name -> structural
    ("FLR-PLN_Floor Plan", "04", "new_New", "ar-Architecture",
     "FLR-PLN_04_new_ar", "FLR-PLN_04_new_ar"),
    # 6 the two generated-name columns disagree -> structural
    ("FLR-PLN_Floor Plan", "04", "new_New", "ar-Architecture",
     "FLR-PLN_04_new_ar", "FLR-PLN_04_new_fa"),
    # 7 empty composite field (level missing) -> structural
    ("FLR-PLN_Floor Plan", "", "new_New", "ar-Architecture",
     "FLR-PLN__new_ar", "FLR-PLN__new_ar"),
    # 8 level 'L05' does not resolve -> referential orphan
    ("FLR-PLN_Floor Plan", "L05", "new_New", "ar-Architecture",
     "FLR-PLN_L05_new_ar", "FLR-PLN_L05_new_ar"),
]


def _build_workbook(path):
    wb = openpyxl.Workbook()
    ws17 = wb.active
    ws17.title = "17-Fixture CNN Data"
    for header_col, code_col, header, (sub_a, sub_b), rows in _BLOCKS:
        ws17[f"{header_col}2"] = header
        ws17[f"{header_col}3"] = sub_a
        if code_col and sub_b:
            ws17[f"{code_col}3"] = sub_b
        r = 5
        for value, code in rows:
            ws17[f"{header_col}{r}"] = value
            if code_col and code is not None:
                ws17[f"{code_col}{r}"] = code
            r += 1

    wsv = wb.create_sheet("05-Views")
    for col, label in _VIEWS_HEADERS.items():
        wsv[f"{col}1"] = label
    r = 2
    for vt, lvl, ph, disc, gen, val in _VIEWS_ROWS:
        wsv[f"B{r}"] = vt
        wsv[f"C{r}"] = lvl
        wsv[f"D{r}"] = ph
        wsv[f"F{r}"] = disc
        wsv[f"H{r}"] = gen
        wsv[f"I{r}"] = val
        r += 1
    wb.save(path)


@pytest.fixture
def workbook(tmp_path):
    p = tmp_path / "ccn_rules_fixture.xlsx"
    _build_workbook(str(p))
    return str(p)


# ----------------------- Deliverable (a): row reader ----------------------

def test_read_views_rows_returns_one_row_per_data_row(workbook):
    rows = ccn_views.read_views_rows(workbook)
    # 8 seeded data rows (the header row and blanks are not data)
    assert len(rows) == 8


def test_row_fields_resolved_by_header_text(workbook):
    rows = ccn_views.read_views_rows(workbook)
    first = rows[0]
    assert first.generated_name == "FLR-PLN_04_new_ar"
    assert first.value_only == "FLR-PLN_04_new_ar"
    assert first.view_type == "FLR-PLN_Floor Plan"
    assert first.level == "04"
    assert first.phase == "new_New"
    assert first.discipline == "ar-Architecture"


def test_row_carries_one_indexed_source_row(workbook):
    rows = ccn_views.read_views_rows(workbook)
    # data starts at Excel row 2 (header on row 1)
    assert rows[0].row == 2
    assert rows[-1].row == 9


def test_broken_formula_row_is_still_read(workbook):
    rows = ccn_views.read_views_rows(workbook)
    broken = [r for r in rows if r.generated_name == "#REF!"]
    assert len(broken) == 1
    assert broken[0].row == 5


# ----------------------- Deliverable (b+c): rules engine ------------------
import os

from agent.ccn import rules as ccn_rules


def _counts(workbook):
    return ccn_rules.summarize(ccn_rules.run_rules(workbook))["by_rule"]


def test_structural_counts_on_fixture(workbook):
    c = _counts(workbook)
    assert c.get("broken_formula") == 1          # the #REF! row
    assert c.get("generated_name_mismatch") == 1  # the disagreeing-columns row
    assert c.get("duplicate_name") == 1           # FLR-PLN_04_new_ar reused
    assert c.get("empty_required_field") == 1     # the empty-level row
    # no view-type abbreviation collision is seeded
    assert c.get("abbrev_collision", 0) == 0


def test_referential_orphans_grouped_with_nearest_match(workbook):
    findings = ccn_rules.run_rules(workbook)
    orphans = [f for f in findings if f.rule_id == "orphan_token"]
    by_value = {f.offending_value: f for f in orphans}
    # 'ex' (Construction Nature) and 'L05' (Levels) don't resolve
    assert "ex" in by_value and "L05" in by_value
    # grouped: one finding for 'ex' though it appears on two rows
    assert "used on 2 row" in by_value["ex"].message
    # nearest-match is case-insensitive: ex -> EXT
    assert by_value["ex"].nearest_match == "EXT"
    # a resolving value (phase 'new' -> NEW) produces no orphan
    assert "new" not in by_value


def test_findings_are_cell_traceable(workbook):
    findings = ccn_rules.run_rules(workbook)
    assert findings, "expected findings on the seeded fixture"
    for f in findings:
        assert f.sheet == "05-Views"
        assert f.cell and f.cell[0].isalpha() and f.cell[1:].isdigit()
        assert f.severity in {"high", "medium", "low"}


def test_duplicate_finding_reports_all_rows(workbook):
    findings = ccn_rules.run_rules(workbook)
    dup = [f for f in findings if f.rule_id == "duplicate_name"]
    assert len(dup) == 1
    # FLR-PLN_04_new_ar is on data rows 2, 6, 7 (Excel-1-indexed)
    assert "used on 3 rows" in dup[0].message


def test_abbrev_collision_detects_view_type_mismatch(tmp_path):
    """A generated name whose view-type token disagrees with the declared one."""
    p = tmp_path / "collide.xlsx"
    _build_workbook(str(p))
    wb = openpyxl.load_workbook(str(p))
    wsv = wb["05-Views"]
    # row 2: declared view type FLR-PLN, but generated name leads with WL-PLN
    wsv["H2"] = "WL-PLN_04_new_ar"
    wsv["I2"] = "WL-PLN_04_new_ar"
    wb.save(str(p))
    findings = ccn_rules.run_rules(str(p))
    collisions = [f for f in findings if f.rule_id == "abbrev_collision"]
    assert len(collisions) == 1
    assert collisions[0].offending_value == "WL-PLN"
    assert collisions[0].nearest_match == "FLR-PLN"


_REAL_WB = ("/Users/ayaan/Projects/AI_Chatbot/data/"
            "04_Templates_Coding_Classification Naming_V2.01_08172026.xlsx")


@pytest.mark.skipif(not os.path.exists(_REAL_WB), reason="real workbook not present")
def test_regression_against_real_workbook():
    """Locks the verified counts on the actual client workbook (V2.01)."""
    c = ccn_rules.summarize(ccn_rules.run_rules(_REAL_WB))["by_rule"]
    assert c["broken_formula"] == 58
    assert c["generated_name_mismatch"] == 51
    assert c["duplicate_name"] == 22
    assert c["abbrev_collision"] == 1
    assert c["orphan_token"] == 3
