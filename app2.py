# file: app2.py
# BIM Classification Agent — XLSX ingestion + agent-mode chat only

import os
import logging
import traceback
import re
import hashlib
import json
from typing import List, Dict, Any
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
from dotenv import load_dotenv
from openai import OpenAI
import chromadb
from chromadb.config import Settings
from werkzeug.exceptions import RequestEntityTooLarge
import bleach
import time

from agent import AGENT_TOOLS, AGENT_FUNCTION_MAP, DESTRUCTIVE_ACTIONS
from agent import create_modules_blueprint, handle_confirmation, handle_tool_call
from agent.chromadb_sync import init as init_agent_chromadb
from agent.db import init_db as init_classifications_db

# ========== LOGGING SETUP ==========
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# ========== ENVIRONMENT ==========
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    logger.error("OPENAI_API_KEY not found in environment variables")
    raise ValueError("OPENAI_API_KEY is required")


# ========== CONFIGURATION ==========
class Config:
    MAX_CONTENT_LENGTH = 50 * 1024 * 1024  # 50MB max file size
    CHUNK_SIZE = 1200
    CHUNK_OVERLAP = 150
    MAX_RETRIES = 5
    RETRY_DELAY = 2
    EMBEDDING_MODEL = "text-embedding-3-small"
    CHAT_MODEL = "gpt-4o"
    MAX_CONTEXT_CHUNKS = 10
    DEEPSEEK_CHAT_MODEL = "deepseek-chat"


# ========== INIT ==========
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = Config.MAX_CONTENT_LENGTH

# CORS Configuration
if os.getenv("FLASK_ENV") == "development":
    CORS(
        app,
        resources={
            r"/api/*": {
                "origins": "*",
                "methods": ["GET", "POST", "DELETE", "PUT", "OPTIONS"],
                "allow_headers": ["Content-Type"],
                "supports_credentials": False,
            }
        },
    )
else:
    allowed_origins = os.getenv(
        "CORS_ORIGINS", "http://localhost:3000,http://localhost:5176"
    ).split(",")
    CORS(
        app,
        resources={
            r"/api/*": {
                "origins": allowed_origins,
                "methods": ["GET", "POST", "DELETE", "PUT", "OPTIONS"],
                "allow_headers": ["Content-Type"],
            }
        },
    )

# Rate Limiting
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# Caching Configuration - tries Redis, falls back to simple memory cache
redis_url = os.getenv("REDIS_URL", "redis://redis:6379/0")
try:
    cache = Cache(
        app,
        config={
            "CACHE_TYPE": "redis",
            "CACHE_REDIS_URL": redis_url,
            "CACHE_DEFAULT_TIMEOUT": 300,
        },
    )
    logger.info(f"Connected to Redis cache at {redis_url}")
except Exception as e:
    logger.warning(f"Redis cache unavailable ({e}), using simple memory cache")
    cache = Cache(app, config={"CACHE_TYPE": "SimpleCache"})

openai_client = OpenAI(api_key=OPENAI_API_KEY, max_retries=3, timeout=30.0)

DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
deepseek_client = None
if DEEPSEEK_API_KEY:
    deepseek_client = OpenAI(
        api_key=DEEPSEEK_API_KEY,
        base_url="https://api.deepseek.com",
        max_retries=3,
        timeout=60.0,
    )
    logger.info("DeepSeek client initialized for chat")
else:
    logger.warning("DEEPSEEK_API_KEY not found — using OpenAI for chat")


# Security headers middleware
@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = (
        "max-age=31536000; includeSubDomains"
    )
    return response


# Initialize Chroma with persistent storage
try:
    chroma = chromadb.PersistentClient(path="./chroma_db")
    app.config["CHROMADB_IN_MEMORY"] = False
    logger.info("Connected to persistent ChromaDB")
except Exception as e:
    if os.getenv("FLASK_ENV", "production") == "production":
        raise RuntimeError(f"ChromaDB init failed in production: {e}") from e
    logger.warning(f"Using in-memory ChromaDB (dev/test only): {e}")
    chroma = chromadb.Client(Settings())
    app.config["CHROMADB_IN_MEMORY"] = True

# Create or load Chroma collection
try:
    collection = chroma.get_collection("docs")
    logger.info("Loaded existing collection 'docs'")
except Exception:
    collection = chroma.create_collection("docs")
    logger.info("Created new collection 'docs'")


# ========== UTILS ==========
def sanitize_input(text: str, max_length: int = 5000) -> str:
    """Sanitize user input to prevent injection attacks"""
    if not text:
        return ""
    text = bleach.clean(text, tags=[], strip=True)
    text = text[:max_length]
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]", "", text)
    return text.strip()


def validate_file_path(file_path: str) -> bool:
    """Validate file path exists and is an XLSX file"""
    if not file_path:
        return False
    file_path = os.path.normpath(file_path)
    if ".." in file_path:
        return False
    if not os.path.exists(file_path):
        return False
    file_ext = os.path.splitext(file_path)[1].lower()
    return file_ext == ".xlsx"


def get_embedding_cached(text: str) -> List[float]:
    """Get single embedding with caching"""
    text_hash = hashlib.md5(text.encode("utf-8")).hexdigest()
    cache_key = f"embedding:{text_hash}"

    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    for attempt in range(Config.MAX_RETRIES):
        try:
            response = openai_client.embeddings.create(
                model=Config.EMBEDDING_MODEL, input=[text]
            )
            embedding = response.data[0].embedding
            cache.set(cache_key, embedding, timeout=86400)  # 24h
            return embedding
        except Exception as e:
            logger.warning(f"Embedding attempt {attempt + 1} failed: {e}")
            if attempt < Config.MAX_RETRIES - 1:
                time.sleep(Config.RETRY_DELAY * (attempt + 1))
            else:
                raise


# ========== XLSX INGESTION ==========
def extract_text_from_xlsx(filepath: str) -> List[str]:
    """Extract text from all sheets of an XLSX file. Returns one string per sheet."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        results = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts = [
                f"Sheet: {sheet_name} — {os.path.basename(filepath)}",
                "-" * 80,
            ]

            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    row_text = " | ".join(
                        str(cell).strip() if cell is not None else ""
                        for cell in row
                    )
                    text_parts.append(row_text)

            sheet_text = "\n".join(text_parts)
            if sheet_text.strip():
                results.append(sheet_text)

        if not results:
            raise ValueError("No text could be extracted from XLSX file")

        logger.info(f"Extracted {len(results)} sheet(s) from XLSX: {filepath}")
        return results

    except Exception as e:
        logger.error(f"Failed to extract text from {filepath}: {e}")
        raise


_CSI_CODE_RE = re.compile(r"^\d{2}\s\d{2}")
_UNIFORMAT_CODE_RE = re.compile(r"^[A-Z]?\d+(\.\d+)?$")

# Sheet-specific parsers for _upsert_xlsx_to_sqlite
_SHEET_PARSERS = {}


def _cell(row, idx):
    """Safely get a stripped string from a row tuple."""
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ""


def _parse_uniformat(rows, sheet_name):
    """Parse 06-Uniformat: col A = code, col B = description, col C = level,
    col E = MasterFormat cross-ref, col F = OmniClass cross-ref."""
    entries = []
    for row in rows:
        code_raw = _cell(row, 0)
        if not code_raw or code_raw in ("NUMBER", "UniFormat", "Title",
                                         "Description", "Version", "Function",
                                         "Number Parameter", "Description Parameter"):
            continue
        # Normalise numeric codes (openpyxl may return int/float)
        if isinstance(row[0], (int, float)):
            code_raw = str(row[0]).rstrip("0").rstrip(".") if "." in str(row[0]) else str(int(row[0]))
        if not _UNIFORMAT_CODE_RE.match(code_raw):
            continue
        name = _cell(row, 1)
        if not name:
            continue
        level = _cell(row, 2)
        mf = _cell(row, 4)
        omni = _cell(row, 5)
        desc_parts = []
        if level:
            desc_parts.append(f"Level {level}")
        if mf and mf != "N/A":
            desc_parts.append(f"MasterFormat: {mf}")
        if omni and omni != "N/A":
            desc_parts.append(f"OmniClass: {omni}")
        description = " | ".join(desc_parts)
        # Category: everything before last dot, or letter prefix for top-level
        if "." in code_raw:
            category = code_raw.rsplit(".", 1)[0]
        elif len(code_raw) > 2:
            # e.g. A1010 -> A10, B2010 -> B20
            prefix = ""
            for ch in code_raw:
                if ch.isalpha():
                    prefix += ch
                else:
                    break
            category = prefix + code_raw[len(prefix):-2] if len(code_raw[len(prefix):]) > 2 else prefix
        else:
            category = ""
        entries.append((code_raw, name, description, sheet_name, category))
    return entries


def _parse_families(rows, sheet_name):
    """Parse 03d-Families: rows 4-10 only. Col A = category, B = type,
    C = adjective, D = company, E = generated family name (used as code)."""
    entries = []
    for row in rows[4:11]:  # rows 4 through 10
        if len(row) < 5:
            continue
        family_name = _cell(row, 4)
        category_name = _cell(row, 0)
        if not family_name or not category_name or family_name == "#N/A":
            continue
        type_fn = _cell(row, 1)
        adjective = _cell(row, 2)
        company = _cell(row, 3)
        desc_parts = []
        if type_fn:
            desc_parts.append(f"Type: {type_fn}")
        if adjective:
            desc_parts.append(f"Attributes: {adjective}")
        if company:
            desc_parts.append(f"Company: {company}")
        description = " | ".join(desc_parts)
        entries.append((family_name, category_name, description, sheet_name, category_name))
    return entries


def _parse_bim_filename(rows, sheet_name):
    """Parse 02-BIM FIle Name: row 3 = field headers, rows 4+ = discipline entries.
    Code = BIM-FILE-{discipline}, category = BIM-FILE."""
    entries = []
    # Capture the naming convention from header row (row 3)
    header_fields = []
    if len(rows) > 3:
        for cell in rows[3]:
            if cell is not None:
                # Shorten the verbose header text
                field = str(cell).strip().split("(")[0].strip()
                if field:
                    header_fields.append(field)
    convention = "Fields: " + ", ".join(header_fields) if header_fields else ""

    for row in rows[4:]:
        discipline = _cell(row, 3)
        if not discipline or discipline == "….":
            continue
        company = _cell(row, 2)
        file_type = _cell(row, 11) if len(row) > 11 else ""
        code = f"BIM-FILE-{discipline.strip().rstrip(',')}"
        desc_parts = []
        if company:
            desc_parts.append(f"Company: {company}")
        if file_type:
            desc_parts.append(f"File Type: {file_type}")
        if convention:
            desc_parts.append(convention)
        description = " | ".join(desc_parts)
        entries.append((code, discipline.strip().rstrip(","), description, sheet_name, "BIM-FILE"))
    return entries


def _parse_variable_data(rows, sheet_name):
    """Parse 07-Variable Data: col B = category name, col C = abbreviation (used as code)."""
    entries = []
    for row in rows[1:]:  # skip header row
        category_name = _cell(row, 1)
        abbrev = _cell(row, 2)
        if not category_name or not abbrev:
            continue
        description = f"Revit family abbreviation for {category_name}"
        entries.append((abbrev, category_name, description, sheet_name, "ABBREV"))
    return entries


_SHEET_PARSERS = {
    "06-Uniformat": _parse_uniformat,
    "03d-Families": _parse_families,
    "02-BIM FIle Name": _parse_bim_filename,
    "07-Variable Data": _parse_variable_data,
}


def _upsert_xlsx_to_sqlite(filepath: str) -> int:
    """Parse an XLSX file and upsert classification rows into the SQLite classifications table.

    Handles multiple sheet formats:
    - Normalized file (header: code, name, description)
    - 05-Masterformat (CSI codes in col B)
    - 06-Uniformat (codes in col A with hierarchy levels)
    - 03d-Families (Revit family naming conventions, rows 4-10)
    - 02-BIM FIle Name (BIM file naming template by discipline)
    - 07-Variable Data (Revit category abbreviation lookup)

    Returns the number of rows upserted.
    """
    try:
        import openpyxl
        from agent.db import get_db

        wb = openpyxl.load_workbook(filepath, data_only=True)
        conn = get_db()
        upsert_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            entries = []

            # Check for a dedicated sheet parser first
            if sheet_name in _SHEET_PARSERS:
                entries = _SHEET_PARSERS[sheet_name](rows, sheet_name)
            else:
                # Fallback: normalized format or CSI code detection
                first = next((r for r in rows if any(c is not None for c in r)), None)
                normalized = (
                    first is not None
                    and first[0] is not None
                    and str(first[0]).strip().lower() == "code"
                )

                for row in rows:
                    if len(row) < 2:
                        continue
                    if normalized:
                        if row[0] is not None and str(row[0]).strip().lower() == "code":
                            continue
                        code = _cell(row, 0)
                        name = _cell(row, 1)
                        description = _cell(row, 2)
                    else:
                        if len(row) < 3:
                            continue
                        code = _cell(row, 1)
                        if not _CSI_CODE_RE.match(code):
                            continue
                        name = _cell(row, 2)
                        description = _cell(row, 3)

                    if not code or not name:
                        continue
                    category = code.rsplit(".", 1)[0] if "." in code else ""
                    entries.append((code, name, description, sheet_name, category))

            # Upsert all entries for this sheet
            for code, name, description, sn, category in entries:
                conn.execute(
                    """
                    INSERT INTO classifications
                        (code, name, description, sheet, category, source)
                    VALUES (?, ?, ?, ?, ?, 'xlsx')
                    ON CONFLICT(code) DO UPDATE SET
                        name        = excluded.name,
                        description = excluded.description,
                        sheet       = excluded.sheet,
                        category    = excluded.category,
                        updated_at  = datetime('now')
                    """,
                    (code, name, description, sn, category),
                )
                upsert_count += 1

        conn.commit()
        logger.info(f"Upserted {upsert_count} classifications from {os.path.basename(filepath)}")
        return upsert_count

    except Exception as e:
        logger.warning(f"SQLite upsert failed for {filepath}: {e}")
        return 0


# ========== AGENT WIRING ==========
init_classifications_db()
init_agent_chromadb(collection, get_embedding_cached)
modules_bp = create_modules_blueprint(limiter, sanitize_input)
app.register_blueprint(modules_bp)


# ========== ROUTES ==========
@app.route("/api/health", methods=["GET"])
@limiter.limit("10 per minute")
def health_check():
    """Health check endpoint"""
    try:
        openai_client.models.list()
        collection.count()
        return jsonify(
            {
                "status": "healthy",
                "timestamp": time.time(),
                "collection_count": collection.count(),
                "chromadb_in_memory": app.config.get("CHROMADB_IN_MEMORY", False),
            }
        )
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({"status": "unhealthy", "error": str(e)}), 500


@app.route("/api/ingest", methods=["POST"])
@limiter.limit("10 per hour")
def ingest():
    """Ingest an XLSX classification file"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "JSON payload required"}), 400

        file_path = data.get("file_path")
        if not validate_file_path(file_path):
            return jsonify(
                {"error": "Invalid file path or unsupported format. Only XLSX files are accepted."}
            ), 400

        logger.info(f"Starting XLSX ingestion of {file_path}")

        db_rows = _upsert_xlsx_to_sqlite(file_path)

        logger.info(f"Successfully ingested {os.path.basename(file_path)}: {db_rows} classifications upserted")

        return jsonify(
            {
                "status": "success",
                "message": "XLSX file ingested successfully",
                "file_name": os.path.basename(file_path),
                "classifications_upserted": db_rows,
            }
        )

    except Exception as e:
        logger.error(f"Ingestion failed: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "Ingestion failed", "details": str(e)}), 500


@app.route("/api/chat", methods=["POST"])
@limiter.limit("30 per minute")
def chat():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "JSON payload required"}), 400

        # Handle pending action confirmation
        confirm_action = data.get("confirm_action")
        if confirm_action:
            chat_client = deepseek_client or openai_client
            chat_model = (
                Config.DEEPSEEK_CHAT_MODEL if deepseek_client else Config.CHAT_MODEL
            )
            return handle_confirmation(
                confirm_action, sanitize_input,
                chat_client, chat_model,
                openai_client, Config.CHAT_MODEL,
            )

        user_question = data.get("question")
        if not user_question or not user_question.strip():
            return jsonify({"error": "Question is required"}), 400

        user_question = sanitize_input(user_question)
        logger.info(f"Processing agent question: {user_question[:100]}...")

        # ChromaDB context search (agent mode)
        context = ""
        if collection.count() > 0:
            try:
                query_embedding = get_embedding_cached(user_question)
                agent_results = collection.query(
                    query_embeddings=[query_embedding],
                    n_results=Config.MAX_CONTEXT_CHUNKS,
                    where={"source_type": {"$nin": ["blog"]}},
                    include=["documents", "metadatas", "distances"],
                )
                agent_docs = agent_results.get("documents", [[]])[0]
                context = "\n\n".join(
                    f"Source {i + 1}: {doc}" for i, doc in enumerate(agent_docs)
                )
            except Exception as _e:
                logger.warning(f"ChromaDB search failed: {_e}")

        system_prompt = (
            "You are a classification management assistant for Syntec Group.\n"
            "You manage building classification codes in a BIM coding system (e.g., '04 05 13.A1').\n\n"
            "WORKFLOW:\n"
            "1. Review the document context provided — it shows existing classification entries from ingested files\n"
            "2. Use list_category to get a structured list of entries in a specific code category\n"
            "3. Determine the next available sub-code based on existing entries before adding new ones\n"
            "4. Use add_module to create new entries (they are stored as agent-added entries)\n\n"
            "TOOLS:\n"
            "- list_category: List all entries whose code starts with a prefix (e.g., '04 05 13') — use first to understand existing codes\n"
            "- get_module: Look up a specific entry by name or code\n"
            "- list_modules: Show all entries in the database\n"
            "- add_module: Add a new classification entry\n"
            "- update_module: Change an entry's code or description\n"
            "- delete_module: Remove an entry\n\n"
            "CODE FORMAT: BIM codes like '04 05 13.A1' (section.subcategory)\n"
            "- Use markdown formatting in your responses\n"
            "- Always confirm the action result clearly\n"
        )
        user_prompt = (
            f"[CLASSIFICATION CONTEXT START]\n{context}\n[CLASSIFICATION CONTEXT END]\n\n"
            f"User request: {user_question}"
        )

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

        chat_client = deepseek_client or openai_client
        chat_model = (
            Config.DEEPSEEK_CHAT_MODEL if deepseek_client else Config.CHAT_MODEL
        )

        call_kwargs = {
            "model": chat_model,
            "messages": messages,
            "max_tokens": 3000,
            "temperature": 0.4,
            "tools": AGENT_TOOLS,
            "tool_choice": "auto",
        }

        try:
            response = chat_client.chat.completions.create(**call_kwargs)
        except Exception as e:
            if chat_client != openai_client:
                logger.warning(f"DeepSeek chat failed, falling back to OpenAI: {e}")
                call_kwargs["model"] = Config.CHAT_MODEL
                response = openai_client.chat.completions.create(**call_kwargs)
            else:
                raise

        response_message = response.choices[0].message

        tool_response = handle_tool_call(
            response_message, messages, user_question, sanitize_input,
            chat_client, chat_model, openai_client, Config.CHAT_MODEL,
            deepseek_client, collection.count(),
        )
        if tool_response is not None:
            return tool_response

        answer = response_message.content
        logger.info("Successfully generated agent response")

        return jsonify({"answer": answer, "sources": []})

    except Exception as e:
        logger.error(f"Chat failed: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "Failed to process question", "details": str(e)}), 500


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5001))
    debug = os.getenv("FLASK_ENV") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)
