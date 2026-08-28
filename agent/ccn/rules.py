"""CCN validation rules engine (Step 2) — structural + referential checks.

Runs over ``05-Views`` (via :mod:`agent.ccn.views`) and the governed
vocabularies in ``17-*`` (via :mod:`agent.ccn.parse`), emitting :class:`Finding`
records that each trace to one spreadsheet cell.

Two rule classes:

* **Structural** — Views alone, no vocabulary needed: broken formula results,
  disagreement between the two generated-name columns, duplicate generated
  names, empty required component fields, and abbreviation collisions between a
  generated name and the abbreviations declared in its own row.
* **Referential** — Views tokens vs the sheet-17 vocabularies: every Level /
  Phase-Nature / Discipline token must resolve to an entry in the corresponding
  governed list; unresolved tokens are *orphans*, grouped by value with a row
  count and a nearest-match suggestion.

Rules are declared in a data-driven registry (:data:`RULES`) so new checks are
added as config, not scattered logic. View-type tokens are deliberately **not**
validated referentially — sheet 17 has no view-type vocabulary, so their
correctness is checked structurally (abbreviation collision) instead.
"""
from __future__ import annotations

import difflib
from dataclasses import dataclass
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .parse import VIEWS_SHEET, _load_grid, _norm, _views_header_row, parse_workbook
from .views import _COLUMN_SPECS, read_views_rows

VIEWS = VIEWS_SHEET

# Spreadsheet error literals that indicate a broken formula result.
ERROR_MARKERS = ("#REF!", "#VALUE!", "#N/A", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!", "#####")

# Which generated-name token position maps to which sheet-17 vocabulary.
# Position in the underscore-split generated name: 0=view type, 1=level,
# 2=phase/nature, 3=discipline. View type (0) has no governed vocabulary.
_REFERENTIAL_FIELDS = (
    ("level", "Levels"),
    ("phase", "Construction Nature"),
    ("discipline", "Disciplines"),
)


@dataclass(frozen=True)
class Finding:
    """One validation defect, traceable to a single cell."""

    rule_id: str
    severity: str            # "high" | "medium" | "low"
    sheet: str
    row: int                 # 1-indexed source row
    cell: str                # e.g. "C13"
    offending_value: str
    nearest_match: Optional[str]
    message: str


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _abbr(component: str) -> str:
    """Extract the abbreviation from a declared component cell.

    Components are ``ABBR_Full Name`` (view type, phase) or ``abbr-Full Name``
    (discipline); a bare value (level) is its own abbreviation. The abbreviation
    is the text before the first ``_`` or ``-`` separator that precedes a name.
    """
    s = _norm(component)
    if not s:
        return ""
    # phase/view-type use '_'; discipline uses '-'. Prefer '_' when present.
    if "_" in s:
        return s.split("_", 1)[0]
    if "-" in s:
        # only treat as ABBR-Full when there's an alpha name after the dash
        head, tail = s.split("-", 1)
        if tail and tail[:1].isalpha():
            return head
    return s


def _tokens(generated_name: str) -> list:
    """Split a generated name into its underscore-delimited tokens."""
    return [t for t in _norm(generated_name).split("_")]


def _has_error_marker(value: str) -> bool:
    v = _norm(value)
    return any(m in v for m in ERROR_MARKERS)


def _view_column_letters(grid: list) -> dict:
    """Resolve each Views field to its column *letter* for cell references."""
    header_row = _views_header_row(grid)
    header = grid[header_row] if header_row < len(grid) else ()
    letters: dict = {}
    for c in range(len(header)):
        text = _norm(header[c]).lower()
        if not text:
            continue
        for name, predicate in _COLUMN_SPECS:
            if name not in letters and predicate(text):
                letters[name] = get_column_letter(c + 1)
    return letters


def _acceptable_tokens(vocab) -> set:
    """The set of strings (lowercased) a token may match for one vocabulary."""
    tokens = set()
    for e in vocab.entries:
        if e.value:
            tokens.add(_norm(e.value).lower())
        if e.code:
            tokens.add(_norm(e.code).lower())
    return tokens


def _vocab_candidates(vocab) -> list:
    """Human-facing candidate strings for nearest-match suggestions."""
    out = []
    for e in vocab.entries:
        out.append(e.code or e.value)
    return [c for c in out if c]


def _nearest(value: str, candidates: list) -> Optional[str]:
    """Case-insensitive nearest-match suggestion, returning the original casing."""
    lower_map: dict = {}
    for c in candidates:
        lower_map.setdefault(_norm(c).lower(), c)
    hits = difflib.get_close_matches(_norm(value).lower(), list(lower_map), n=1, cutoff=0.3)
    return lower_map[hits[0]] if hits else None


@dataclass
class _Context:
    """Everything the rule functions need, gathered once."""

    rows: list
    vocabs: dict                       # name -> Vocabulary
    letters: dict                      # field -> column letter
    sheet: str = VIEWS

    def cell(self, field_name: str, row: int) -> str:
        return f"{self.letters.get(field_name, '?')}{row}"


# --------------------------------------------------------------------------- #
# structural rules
# --------------------------------------------------------------------------- #
def rule_broken_formula(ctx: _Context) -> list:
    out = []
    for vr in ctx.rows:
        if _has_error_marker(vr.generated_name):
            out.append(Finding(
                "broken_formula", "high", ctx.sheet, vr.row,
                ctx.cell("generated_name", vr.row), vr.generated_name, None,
                "Generated name holds a broken formula result "
                f"({vr.generated_name!r}); the naming formula did not resolve.",
            ))
    return out


def rule_generated_name_mismatch(ctx: _Context) -> list:
    out = []
    for vr in ctx.rows:
        g, v = _norm(vr.generated_name), _norm(vr.value_only)
        if not g or not v:
            continue
        if _has_error_marker(g) or _has_error_marker(v):
            continue  # broken_formula owns these
        if g != v:
            out.append(Finding(
                "generated_name_mismatch", "high", ctx.sheet, vr.row,
                ctx.cell("value_only", vr.row), vr.value_only, g,
                "The two generated-name columns disagree: "
                f"'View Name, Generated'={g!r} vs 'View Name Value Only'={v!r} "
                "— they must encode the same fields in the same order.",
            ))
    return out


def rule_duplicate_name(ctx: _Context) -> list:
    seen: dict = {}
    for vr in ctx.rows:
        g = _norm(vr.generated_name)
        if not g or _has_error_marker(g):
            continue
        seen.setdefault(g, []).append(vr.row)
    out = []
    for name, rows in seen.items():
        if len(rows) > 1:
            out.append(Finding(
                "duplicate_name", "medium", ctx.sheet, rows[0],
                ctx.cell("generated_name", rows[0]), name, None,
                f"Generated name {name!r} is used on {len(rows)} rows "
                f"({', '.join(map(str, rows[:8]))}"
                f"{'…' if len(rows) > 8 else ''}); the standard requires "
                "unique view names.",
            ))
    return out


def rule_empty_required_field(ctx: _Context) -> list:
    """A required component field empty on a row that otherwise holds data."""
    required = ("view_type", "level", "phase", "discipline")
    out = []
    for vr in ctx.rows:
        present = any(_norm(getattr(vr, f)) for f in required)
        if not present:
            continue
        for f in required:
            if not _norm(getattr(vr, f)):
                out.append(Finding(
                    "empty_required_field", "medium", ctx.sheet, vr.row,
                    ctx.cell(f, vr.row), "", None,
                    f"Required component '{f}' is empty on a populated view row; "
                    "the generated name cannot be composed correctly.",
                ))
    return out


def rule_abbrev_collision(ctx: _Context) -> list:
    """View-type abbreviation in the generated name disagrees with the row's.

    Only the leading view-type token is checked: it is the one positionally
    stable slot (generated names have variable middle structure — enlarged
    plans insert room descriptors — so fixed-index matching of the other
    components would raise false accusations). Level / Phase / Discipline are
    validated instead by the referential ``orphan_token`` rule.
    """
    out = []
    for vr in ctx.rows:
        g = _norm(vr.generated_name)
        if not g or _has_error_marker(g):
            continue
        toks = _tokens(g)
        if not toks:
            continue
        used = toks[0]
        declared = _abbr(vr.view_type)
        if used and declared and used.lower() != declared.lower():
            out.append(Finding(
                "abbrev_collision", "high", ctx.sheet, vr.row,
                ctx.cell("generated_name", vr.row), used, declared,
                f"Generated name's view-type abbreviation {used!r} disagrees "
                f"with the row's declared view type {declared!r}.",
            ))
    return out


# --------------------------------------------------------------------------- #
# referential rules
# --------------------------------------------------------------------------- #
def rule_orphan_tokens(ctx: _Context) -> list:
    """Level / Phase / Discipline values not found in their sheet-17 vocabulary.

    Grouped by (field, value): one finding per distinct orphan value with a row
    count and a nearest-match suggestion, not one line per row.
    """
    out = []
    for field_name, vocab_name in _REFERENTIAL_FIELDS:
        vocab = ctx.vocabs.get(vocab_name)
        if vocab is None:
            continue
        accepted = _acceptable_tokens(vocab)
        candidates = _vocab_candidates(vocab)
        groups: dict = {}
        for vr in ctx.rows:
            raw = getattr(vr, field_name)
            if field_name == "level":
                value = _norm(raw)
            else:
                value = _abbr(raw)
            if not value:
                continue
            if value.lower() in accepted:
                continue
            groups.setdefault(value, []).append(vr.row)
        for value, rows in sorted(groups.items(), key=lambda kv: -len(kv[1])):
            near = _nearest(value, candidates)
            out.append(Finding(
                "orphan_token", "high", ctx.sheet, rows[0],
                ctx.cell(field_name, rows[0]), value, near,
                f"{field_name.capitalize()} value {value!r} is not in the "
                f"governed '{vocab_name}' vocabulary — used on {len(rows)} row(s)"
                + (f"; did you mean {near!r}?" if near else "."),
            ))
    return out


# --------------------------------------------------------------------------- #
# registry
# --------------------------------------------------------------------------- #
RULES = (
    {"id": "broken_formula", "kind": "structural", "fn": rule_broken_formula},
    {"id": "generated_name_mismatch", "kind": "structural", "fn": rule_generated_name_mismatch},
    {"id": "duplicate_name", "kind": "structural", "fn": rule_duplicate_name},
    {"id": "empty_required_field", "kind": "structural", "fn": rule_empty_required_field},
    {"id": "abbrev_collision", "kind": "structural", "fn": rule_abbrev_collision},
    {"id": "orphan_token", "kind": "referential", "fn": rule_orphan_tokens},
)


def build_context(path: str) -> _Context:
    """Gather rows, vocabularies, and column letters from the workbook once."""
    rows = read_views_rows(path)
    parsed = parse_workbook(path)
    vocabs = {v.name: v for v in parsed.vocabularies if v.sheet.startswith("17-")}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        grid = _load_grid(wb[VIEWS]) if VIEWS in wb.sheetnames else []
    finally:
        wb.close()
    return _Context(rows=rows, vocabs=vocabs, letters=_view_column_letters(grid))


def run_rules(path: str, context: Optional[_Context] = None) -> list:
    """Run every rule against the workbook and return all findings."""
    ctx = context or build_context(path)
    findings: list = []
    for rule in RULES:
        findings.extend(rule["fn"](ctx))
    return findings


def summarize(findings: list) -> dict:
    """Counts by rule id and by severity."""
    by_rule: dict = {}
    by_severity: dict = {}
    for f in findings:
        by_rule[f.rule_id] = by_rule.get(f.rule_id, 0) + 1
        by_severity[f.severity] = by_severity.get(f.severity, 0) + 1
    return {"total": len(findings), "by_rule": by_rule, "by_severity": by_severity}


def format_report(findings: list, limit: int = 40) -> str:
    """Human-readable report: summary counts, then findings (capped)."""
    s = summarize(findings)
    lines = ["CCN VALIDATION REPORT", "=" * 60,
             f"{s['total']} findings  |  by severity: {s['by_severity']}",
             "by rule:"]
    for rid, n in sorted(s["by_rule"].items(), key=lambda kv: -kv[1]):
        lines.append(f"    {n:>4}  {rid}")
    lines.append("-" * 60)
    ordered = sorted(findings, key=lambda f: ({"high": 0, "medium": 1, "low": 2}
                                              .get(f.severity, 3), f.rule_id, f.row))
    for f in ordered[:limit]:
        near = f" (did you mean {f.nearest_match!r}?)" if f.nearest_match else ""
        lines.append(f"[{f.severity:<6}] {f.sheet}!{f.cell} {f.rule_id}: "
                     f"{f.offending_value!r}{near}")
        lines.append(f"          {f.message}")
    if len(findings) > limit:
        lines.append(f"... and {len(findings) - limit} more (raise --limit to see all)")
    return "\n".join(lines)


def main(argv=None) -> int:
    import sys
    argv = list(sys.argv[1:] if argv is None else argv)
    if not argv:
        print("usage: python -m agent.ccn.rules <workbook_path> [--limit N]",
              file=sys.stderr)
        return 2
    path = argv[0]
    limit = 40
    if "--limit" in argv:
        try:
            limit = int(argv[argv.index("--limit") + 1])
        except (ValueError, IndexError):
            print("--limit requires an integer", file=sys.stderr)
            return 2
    print(format_report(run_rules(path), limit=limit))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
