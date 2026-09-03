"""Parse the CCN classification workbook into structured vocabularies.

Step 1 of the CCN validator: read the vocabulary/reference data and emit a
parse report. **No validation rules** — those are a separate later task.

Two sheets are in scope:

* ``17-*`` (resolved by prefix) holds eight independent vertical vocab lists
  parked side by side, separated by blank spacer columns. Each block is laid
  out as::

      row 2  SECTION HEADER      (e.g. "REVIT CATEGORIES")
      row 3  column sub-header   (e.g. "Revit Category Name" | "Abbreviation")
      row 4  (blank spacer)
      row 5+ data ................ one entry per row

  Column letters drift between revisions, so blocks are DETECTED by matching
  their section-header text in the header row (leftmost match wins) — never
  hardcoded by position. Each block owns its column, so its entries are simply
  the non-empty cells in that column below the sub-header.

* ``05-Views`` holds a real data table plus two hazards that must be
  skipped-and-counted, never parsed as data: loose guidance *prose* sitting in
  data-looking columns, and a *stale* superseded vocab block whose values carry
  ``(X)`` letter codes and disagree with sheet 17.

Non-data cells are always counted, never silently dropped: the report states
"N cells ignored as non-data".
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

SHEET17_PREFIX = "17-"
VIEWS_SHEET = "05-Views"

# Data-driven block specs. Each block is found by matching ``header`` (anchored,
# case-insensitive) against cells in the detected header row; the leftmost match
# wins, which disambiguates repeated headers such as "FAMILIES" appearing again
# far to the right of the sheet.
BLOCK_SPECS = (
    {"name": "Revit Categories", "header": r"REVIT CATEGOR"},
    {"name": "Uniformat", "header": r"UNIFORMAT"},
    {"name": "Levels", "header": r"LEVELS"},
    {"name": "Construction Nature", "header": r"CONSTRUCTION NATURE"},
    {"name": "Disciplines", "header": r"DISCIPLINES"},
    {"name": "Families", "header": r"FAMILIES"},
    {"name": "Organizations", "header": r"ORGANIZATIONS"},
    {"name": "Sheet Classifications", "header": r"SHEETS?"},
)

# A cell is "prose" (guidance, not data) when it is a longish sentence.
_PROSE_MIN_LEN = 60
# A "stale" column restates vocab with parenthetical letter codes, e.g.
# "Civil (C)", "Demo (D)", "Architecture (A)".
_PAREN_CODE_RE = re.compile(r"\([A-Za-z0-9]{1,3}\)\s*$")


@dataclass
class Entry:
    """One vocabulary entry: its primary value plus an optional adjacent code."""

    value: str
    code: Optional[str]
    row: int


@dataclass
class Vocabulary:
    """A single named vocabulary list detected on a sheet."""

    name: str
    sheet: str
    entries: list = field(default_factory=list)

    @property
    def count(self) -> int:
        return len(self.entries)

    @property
    def first(self) -> Optional[str]:
        return self.entries[0].value if self.entries else None

    @property
    def last(self) -> Optional[str]:
        return self.entries[-1].value if self.entries else None


@dataclass
class ParseResult:
    """Everything the parser produced from one workbook."""

    vocabularies: list = field(default_factory=list)
    ignored_cells: int = 0
    ignored_breakdown: dict = field(default_factory=dict)


def _norm(value) -> str:
    return "" if value is None else str(value).strip()


def _load_grid(ws) -> list:
    """Read a worksheet into a list of row tuples (values only, once)."""
    return list(ws.iter_rows(values_only=True))


def _cell(grid: list, row: int, col: int) -> str:
    """0-indexed, bounds-safe accessor returning the stripped string value."""
    if 0 <= row < len(grid):
        r = grid[row]
        if 0 <= col < len(r):
            return _norm(r[col])
    return ""


def _find_header_row(grid: list) -> int:
    """Return the 0-indexed row where the most section headers coincide."""
    patterns = [re.compile(spec["header"], re.IGNORECASE) for spec in BLOCK_SPECS]
    best_row, best_hits = 0, -1
    for r, row in enumerate(grid):
        hits = 0
        for c in range(len(row)):
            text = _norm(row[c])
            if text and any(p.match(text) for p in patterns):
                hits += 1
        if hits > best_hits:
            best_row, best_hits = r, hits
    return best_row


def _detect_block_columns(grid: list, header_row: int) -> dict:
    """Map each block name -> its 0-indexed header column (leftmost match)."""
    found = {}
    header_cells = grid[header_row] if header_row < len(grid) else ()
    for spec in BLOCK_SPECS:
        pattern = re.compile(spec["header"], re.IGNORECASE)
        for c in range(len(header_cells)):
            if pattern.match(_norm(header_cells[c])):
                if spec["name"] not in found:  # leftmost wins
                    found[spec["name"]] = c
                break
    return found


def parse_sheet17(grid: list, sheet_name: str):
    """Parse the eight side-by-side vocab blocks. Returns (vocabs, ignored)."""
    header_row = _find_header_row(grid)
    sub_row = header_row + 1
    data_start = header_row + 3  # skip section header, sub-header and blank spacer
    columns = _detect_block_columns(grid, header_row)

    vocabs: list = []
    ignored = 0
    for spec in BLOCK_SPECS:
        col = columns.get(spec["name"])
        if col is None:
            continue
        code_col = col + 1
        # Does the adjacent column carry a code sub-header for this block?
        has_code = bool(_cell(grid, sub_row, code_col))

        vocab = Vocabulary(name=spec["name"], sheet=sheet_name)
        for r in range(data_start, len(grid)):
            value = _cell(grid, r, col)
            if not value:
                continue
            code = _cell(grid, r, code_col) if has_code else ""
            vocab.entries.append(Entry(value=value, code=code or None, row=r + 1))
        vocabs.append(vocab)

        # Count the section-header + sub-header cells we deliberately skipped.
        for header_r in (header_row, sub_row):
            for cc in (col, code_col):
                if _cell(grid, header_r, cc):
                    ignored += 1
    return vocabs, ignored


def _views_header_row(grid: list) -> int:
    """Find the Views table header row (most short-label columns in a row)."""
    labels = ("view type", "abbreviation", "discipline", "level", "phase",
              "sheet", "adjective")
    best_row, best_hits = 0, -1
    for r, row in enumerate(grid):
        hits = 0
        for c in range(len(row)):
            text = _norm(row[c]).lower()
            if text and len(text) < 40 and any(lbl in text for lbl in labels):
                hits += 1
        if hits > best_hits:
            best_row, best_hits = r, hits
    return best_row


def _stale_columns(grid: list, header_row: int) -> set:
    """Columns whose data is dominated by ``(X)`` codes -> superseded block."""
    ncol = max((len(r) for r in grid), default=0)
    stale = set()
    for c in range(ncol):
        hits = sum(
            1
            for r in range(header_row + 1, len(grid))
            if _PAREN_CODE_RE.search(_cell(grid, r, c))
        )
        if hits >= 2:
            stale.add(c)
    return stale


def parse_views(grid: list):
    """Parse 05-Views: expose the canonical Views vocab, count prose + stale.

    Returns (vocabs, ignored, breakdown_detail).
    """
    header_row = _views_header_row(grid)
    stale_cols = _stale_columns(grid, header_row)

    # Canonical Views vocabulary: the "Abbreviation_View Type" column.
    vocabs: list = []
    header_cells = grid[header_row] if header_row < len(grid) else ()
    vocab_col = None
    for c in range(len(header_cells)):
        text = _norm(header_cells[c]).lower()
        if "abbreviation" in text and "view type" in text:
            vocab_col = c
            break
    if vocab_col is not None and vocab_col not in stale_cols:
        vocab = Vocabulary(name="Views", sheet=VIEWS_SHEET)
        for r in range(header_row + 1, len(grid)):
            value = _cell(grid, r, vocab_col)
            if value and len(value) < _PROSE_MIN_LEN:
                vocab.entries.append(Entry(value=value, code=None, row=r + 1))
        vocabs.append(vocab)

    # Ignored non-data: stale-block cells + prose cells (outside stale columns).
    stale_count = 0
    prose_count = 0
    for r in range(len(grid)):
        for c in range(len(grid[r])):
            text = _cell(grid, r, c)
            if not text:
                continue
            if c in stale_cols:
                stale_count += 1
            elif len(text) >= _PROSE_MIN_LEN:
                prose_count += 1
    ignored = stale_count + prose_count
    detail = {"prose": prose_count, "stale": stale_count}
    return vocabs, ignored, detail


def parse_workbook(path: str) -> ParseResult:
    """Open the workbook and parse both in-scope sheets."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = ParseResult()
    try:
        sheet17 = next(
            (n for n in wb.sheetnames if n.startswith(SHEET17_PREFIX)), None
        )
        if sheet17:
            grid = _load_grid(wb[sheet17])
            vocabs, ignored = parse_sheet17(grid, sheet17)
            result.vocabularies.extend(vocabs)
            result.ignored_cells += ignored
            result.ignored_breakdown["17"] = ignored

        if VIEWS_SHEET in wb.sheetnames:
            grid = _load_grid(wb[VIEWS_SHEET])
            vocabs, ignored, detail = parse_views(grid)
            result.vocabularies.extend(vocabs)
            result.ignored_cells += ignored
            result.ignored_breakdown[VIEWS_SHEET] = ignored
            result.ignored_breakdown["05-Views-detail"] = detail
    finally:
        wb.close()
    return result


def format_report(result: ParseResult) -> str:
    """Render the parse report: every vocab (name/count/first/last) + ignored."""
    lines = ["CCN PARSE REPORT", "=" * 60]
    for v in result.vocabularies:
        lines.append(f"{v.name}  [{v.sheet}]")
        lines.append(f"    count: {v.count}")
        lines.append(f"    first: {v.first!r}")
        lines.append(f"    last:  {v.last!r}")
    lines.append("-" * 60)
    lines.append(f"{result.ignored_cells} cells ignored as non-data")
    for key, val in result.ignored_breakdown.items():
        lines.append(f"    {key}: {val}")
    return "\n".join(lines)


def main(argv=None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if not argv:
        print("usage: python -m agent.ccn.parse <workbook_path>", file=sys.stderr)
        return 2
    result = parse_workbook(argv[0])
    print(format_report(result))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
